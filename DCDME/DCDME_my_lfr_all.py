import networkx as nx
from sklearn.metrics.cluster import normalized_mutual_info_score
from sklearn.metrics.cluster import adjusted_rand_score
from collections import defaultdict
import time
import datetime
from sklearn import metrics
import math
import matplotlib.pyplot as plt  # For drawing pictures
import numpy as np
from numpy import linalg as LA
import openpyxl


# global comm_list
# comm_list=[]


def str_to_int(x):
    return [[int(v) for v in line.split()] for line in x]

def node_addition(G, addnodes, communitys):  # The community format entered is {node: community name}
    change_comm = set()  # Community labels that may be found to be changed in the storage structure
    processed_edges = set()  # For processed edges, you need to delete the processed edges from the added edges.

    for u in addnodes:
        neighbors_u = G.neighbors(u)
        neig_comm = set()  # Neighbor's community label
        pc = set()
        for v in neighbors_u:
            if v in communitys:
                neig_comm.add(communitys[v])
            pc.add((u, v))
            pc.add((v, u))  # There is one side in the undirectional map, and it is convenient to add it twice.
        if len(neig_comm) > 1:  # It shows that this joining node is not within the community.
            change_comm = change_comm | neig_comm
            lab = max(communitys.values()) + 1
            communitys.setdefault(u, lab)  # Assign a community tag to u
            change_comm.add(lab)
        else:
            if len(neig_comm) == 1:  # Explain that the node is within the community, or only connected to one community.
                communitys.setdefault(u, list(neig_comm)[0])  # Add nodes to the community
                processed_edges = processed_edges | pc
            else:
                communitys.setdefault(u,
                                      max(communitys.values()) + 1)  # The new node is not connected to other nodes, and a new community label is assigned.

    return change_comm, processed_edges, communitys  # Return to communities that may change, processed edges and the latest community structure.


def node_deletion(G, delnodes, communitys):  # tested, correct
    change_comm = set()  # Community labels that may be found to be changed in the storage structure
    processed_edges = set()  # For the processed side, you need to delete the processed side from the deleted side.
    for u in delnodes:
        neighbors_u = G.neighbors(u)
        neig_comm = set()  # Neighbor's community label
        for v in neighbors_u:
            if v in communitys:
                neig_comm.add(communitys[v])
            processed_edges.add((u, v))
            processed_edges.add((v, u))
        del communitys[u]  # Delete nodes and communities
        change_comm = change_comm | neig_comm
    return change_comm, processed_edges, communitys  # Return to communities that may change, processed edges and the latest community structure.


def edge_addition(addedges, communitys):  # If the joining side is within the community, it will not cause changes in the community, it will not be dealt with, otherwise it will be marked.
    change_comm = set()  # Community labels that may be found to be changed in the storage structure
    #    print addedges
    #    print communitys
    for item in addedges:
        neig_comm = set()  # Neighbor's community label
        neig_comm.add(communitys[item[0]])  # Determine the community where the nodes at both ends are located.
        neig_comm.add(communitys[item[1]])
        if len(neig_comm) > 1:  # It means that this joining is not within the community.
            change_comm = change_comm | neig_comm
    return change_comm  # Return to communities that may change,


def edge_deletion(deledges, communitys):  # If the deleted edge may cause community change within the community, it will not change outside the community.
    change_comm = set()  # Community labels that may be found to be changed in the storage structure
    for item in deledges:
        neig_comm = set()  # Neighbor's community label
        neig_comm.add(communitys[item[0]])  # Determine the community where the nodes at both ends are located.
        neig_comm.add(communitys[item[1]])
        if len(neig_comm) == 1:  # It means that this joining is not within the community.
            change_comm = change_comm | neig_comm

    return change_comm  # Return to communities that may change


def getchangegraph(all_change_comm, newcomm, Gt):
    Gte = nx.Graph()
    com_key = newcomm.keys()
    for v in Gt.nodes():
        if v not in com_key or newcomm[v] in all_change_comm:
            Gte.add_node(v)
            neig_v = Gt.neighbors(v)
            for u in neig_v:
                if u not in com_key or newcomm[u] in all_change_comm:
                    Gte.add_edge(v, u)
                    Gte.add_node(u)

    return Gte


# cdme Community testing****************************************************************

nodecount_comm = defaultdict(int)  # Global variables


def CDME(G):
    deg = G.degree()

    def AA(NA, NB):
        comm_nodes = list(NA & NB)
        sim = 0
        for node in comm_nodes:
            degnode = deg[node]
            if deg[node] == 1:
                degnode = 1.1
            sim = sim + (1.0 / math.log(degnode))
        return sim
        # Compute the jaccard similarity coefficient of two node

    def simjkd(u, v):
        set_v = set(G.neighbors(v))
        set_v.add(v)
        set_u = set(G.neighbors(u))
        set_u.add(u)
        jac = len(set_v & set_u) * 1.0 / len(set_v | set_u)
        return jac

    # Initialize communities of each node
    # In the first stage, a node is a community.
    node_community = dict(zip(G.nodes(), G.nodes()))
    # Compute the core groups
    # The second stage, calculate the core grouping
    st = {}  # storge the AA
    # compute the AA
    for node in G.nodes():
        Nv = sorted(G.neighbors(node))
        for u in Nv:
            Nu = G.neighbors(u)
            keys = str(node) + '_' + str(u)
            st.setdefault(keys, AA(set(Nv), set(Nu)))
    print('AAindex,done')
    for node in G.nodes():
        # The degree of each node
        deg_node = deg[node]
        flag = True
        maxsimdeg = 0
        selected = node
        if deg_node == 1:
            # node_community[node] =  node_community[ G.neighbors(node)[0]]
            node_community[node] = node_community[list(G.neighbors(node))[0]]
        else:
            for neig in G.neighbors(node):
                deg_neig = deg[neig]
                if flag is True and deg_node <= deg_neig:
                    flag = False
                    break

            if flag is False:
                for neig in sorted(G.neighbors(node)):
                    deg_neig = deg[neig]
                    # Compute the Jaccard similarity coefficient
                    # nodesim =  simjkd(node, neig)
                    # Use the AAindex
                    keys = str(node) + '_' + str(neig)
                    nodesim = st[keys]
                    # Compute the node attraction
                    nodesimdeg = deg_neig * nodesim
                    if nodesimdeg > maxsimdeg:
                        selected = neig
                        maxsimdeg = nodesimdeg
                    node_community[node] = node_community[selected]
    old_persum = -(2 ** 63 - 1)
    old_netw_per = -(2 ** 63 - 1)

    persum = old_persum + 1
    netw_per = old_netw_per + 0.1
    maxit = 5
    itern = 0

    print("loop begin:")
    while itern < maxit:
        itern += 1
        old_netw_per = netw_per
        old_persum = persum
        persum = 0
        for node in G.nodes():
            neiglist = sorted(G.neighbors(node))
            cur_p = per(G, node, node_community)  #
            nodeneig_comm = nodecount_comm.keys()
            cur_p_neig = 0

            for neig in neiglist:
                cur_p_neig += per(G, neig, node_community)
            #                #Calculate the number of triangles between the current node and the neighbor node
            #                for neig in neiglist:
            #                    tri=set(neiglist)&set(self.G.neighbors(neig))
            #                    cur_tri += tri
            #                #Degree plus 2 times triangle
            #                cur_p_neig+=2*cur_tri
            for neig_comm in nodeneig_comm:

                node_pre_comm = node_community[node]
                new_p_neig = 0
                if node_pre_comm != neig_comm:
                    node_community[node] = neig_comm
                    new_p = per(G, node, node_community)

                    if cur_p <= new_p:

                        if cur_p == new_p:
                            for newneig in neiglist:
                                new_p_neig += per(G, newneig, node_community)
                            if cur_p_neig < new_p_neig:
                                cur_p = new_p
                                cur_p_neig = new_p_neig
                            else:
                                node_community[node] = node_pre_comm

                        else:
                            for newneig in neiglist:
                                new_p_neig += per(G, newneig, node_community)
                            cur_p = new_p
                            cur_p_neig = new_p_neig
                    else:
                        node_community[node] = node_pre_comm
            persum += cur_p
    #            print(node_community)
    print("loop done")
    # Convert community form, {noble: community} to {community: [noble]}
    #    graph_result = defaultdict(list)
    #    for item in node_community.keys():
    #        node_comm = int(node_community[item])
    #        graph_result[node_comm].append(item)

    return node_community


# The internal degree of node v in a community
def per(G, v, node_community):
    neiglist1 = G.neighbors(v)
    in_v = 0
    # tri=0
    global nodecount_comm

    for neig in neiglist1:
        if node_community[neig] == node_community[v]:
            in_v += 1
            # tri+=len(set(neiglist1)&set(self.G.neighbors(neig)))
        else:

            nodecount_comm[node_community[neig]] += 1

    cin_v = 1.0 * (in_v * in_v)
    per = cin_v
    return per


# ****************************************************************************

def Errorrate(clusters, classes, n):
    # Calculate the error rate, the formula ||A*A'-C*C'||, A' represents the transpose of the matrix. A stands for the community found, the behavior node, listed as a community, and the node belongs to a community, then the corresponding cross cell is 1, otherwise it is 0, and C is the matrix of the real community.
    A = np.zeros((n, len(clusters)), int)
    C = np.zeros((n, len(classes)), int)
    k = 0
    for nodelist in clusters:
        for node in nodelist:
            A[node - 1][k] = 1
        k = k + 1
    k = 0
    for nodelist in classes:
        for node in nodelist:
            C[node - 1][k] = 1
        k = k + 1
    t = A.dot(A.T) - C.dot(C.T)
    errors = LA.norm(t)
    return errors


def purity_score(clusters, classes):
    A = np.c_[(clusters, classes)]

    n_accurate = 0.

    for j in np.unique(A[:, 0]):
        z = A[A[:, 0] == j, 1]
        x = np.argmax(np.bincount(z))
        n_accurate += len(z[z == x])

    return n_accurate / A.shape[0]


def conver_comm_to_lab(comm1):  # Convert the community format to, the label is the main key, and the node is value.
    overl_community = {}
    for node_v, com_lab in comm1.items():
        if com_lab in overl_community.keys():
            overl_community[com_lab].append(node_v)
        else:
            overl_community.update({com_lab: [node_v]})
    return overl_community


def getscore(comm_true, comm_dete, num):
    actual = []
    baseline = []
    for j in range(len(comm_true)):  # groundtruth，j Representing each community, j is the name of the community.
        for c in comm_true[j]:  # Each node in the community represents each node.
            flag = False
            for k in range(len(comm_dete)):  # Detected community, k is the name of the community
                if c in comm_dete[k] and flag == False:
                    flag = True
                    actual.append(j)
                    baseline.append(k)
                    break

    NMI1 = metrics.normalized_mutual_info_score(actual, baseline)
    ARI1 = metrics.adjusted_rand_score(actual, baseline)
    Purity1 = purity_score(baseline, actual)
    errors = 0
    # errors = Errorrate(comm_dete,comm_true,num)
    print('nmi', NMI1)
    print('ari', ARI1)
    print('purity', Purity1)
    print('rate error', errors)

    return NMI1, ARI1, Purity1, errors


def drawcommunity(g, partition, filepath):
    pos = nx.spring_layout(g)
    count1 = 0
    t = 0
    node_color = []
    with open('./colours.txt') as f:
        node_color = f.read().splitlines()
        f.close()


    for com in set(partition.values()):
        count1 = count1 + 1.
        list_nodes = [nodes for nodes in partition.keys()
                      if partition[nodes] == com]

        nx.draw_networkx_nodes(g, pos, list_nodes, node_size=220,
                               node_color=node_color[t])
        nx.draw_networkx_labels(g, pos)
        t = t + 1

    nx.draw_networkx_edges(g, pos, label=True, alpha=0.5)
    plt.savefig(filepath)
    # plt.show()


############################################################
# ----------main-----------------
edges_added = set()
edges_removed = set()
nodes_added = set()
nodes_removed = set()
G = nx.Graph()
allpath = './data/my_LFR/files.txt'
with open(allpath, 'r') as f:
    allpathlist = f.readlines()
    f.close()
# allpathlists=allpathlist[0].strip('\n')
pathfile = ''
for pt in allpathlist:
    pathfile = pt.strip('\n')
    print(pathfile)
    path = './data/my_LFR/' + pathfile + '/'
    edge_file = ''
    comm_file = ''
    G.clear()
    # read edgefile list, where storage the filename of each snapshot
    edgefilelist = []
    commfilelist = []
    with open(path + 'edgeslist.txt', 'r') as f:
        edgefilelist = f.readlines()
        f.close()
    edge_file = edgefilelist[0].strip('\n')
    with open(path + 'commlist.txt', 'r') as f:
        commfilelist = f.readlines()
        f.close()
    comm_file = commfilelist[0].strip('\n')

    # path='./LFR/t/'
    # path='./data/test/'
    with open(path + edge_file, 'r') as f:

        edge_list = f.readlines()
        for edge in edge_list:
            edge = edge.split()
            G.add_node(int(edge[0]))
            G.add_node(int(edge[1]))
            G.add_edge(int(edge[0]), int(edge[1]))
        f.close()
    G = G.to_undirected()
    #Initial diagram
    print('T0 Time film network G0*********************************************')
    nx.draw_networkx(G)
    fpath='./data/pic/'+ pathfile +'_G_0.png'
    plt.savefig(fpath)           #Output method 1: Save the image as a picture file in png format
    plt.show()
    # print G.edges()
    # comm_file='switch.t01.comm'
    nodenumber = G.number_of_nodes()
    with open(path + comm_file, 'r') as f:
        comm_list = f.readlines()
        comm_list = str_to_int(comm_list)  # 真实社区
        f.close()

    # The third stage, the simulation of the Matthew effect stage

    comm = {}  # Used to store the detected community structure, format {node: community tag}
    comm = CDME(G)  # Initial community
    # Painting community
    print('T0 Time Film Community C0*********************************************')
    drawcommunity(G,comm,'./data/pic/' + pathfile + '_community_0.png')
    initcomm = conver_comm_to_lab(comm)
    comm_va = list(initcomm.values())
    commu_num = len(comm_va)
    tru_num = len(comm_list)
    NMI, ARI, Purity, Errors = getscore(comm_list, comm_va, nodenumber)
    path_score = 'result_score_LFR.xlsx'
    # f = open(path_score,'a+')        #Write to the file
    # f.write('path'+"\t"+'NMI'+"\t"+'ARI'+"\t"+'Purity'+'\t'+'detected_community_number'+'ture_community_number'+'errors'"\n" )
    # f.write(path+'_1'+"\t"+str(NMI)+"\t"+str(ARI)+"\t"+str(Purity)+'\t'+str(commu_num)+'\t'+str(tru_num)+str(Errors)+"\n" )
    # f.close()
    # wb=openpyxl.Workbook(path_score)
    # wb.save(path_score)
    wb = openpyxl.load_workbook(filename=path_score)
    ws = wb.create_sheet(path[14:len(path) - 1])
    row = ['path', 'NMI', 'ARI', 'Purity', 'detected_community_number', 'ture_community_number', 'errors']
    ws.append(row)
    row = ['1', str(NMI), str(ARI), str(Purity), str(commu_num), str(tru_num), str(Errors)]
    ws.append(row)

    start = time.time()
    G1 = nx.Graph()
    G2 = nx.Graph()
    G1 = G

    l = len(edgefilelist)
    for i in range(1, l):
        print('begin loop:', i)
        comm_new_file = open(path + commfilelist[i].strip('\n'), 'r')
        comm_new = comm_new_file.readlines()
        comm_new_file.close()
        comm_new = str_to_int(comm_new)

        edge_list_new_file = open(path + edgefilelist[i].strip('\n'), 'r')
        edge_list_new = edge_list_new_file.readlines()
        edge_list_new_file.close()

        #    for line in edge_list_old:
        #         temp = line.strip().split()
        #
        #         G1.add_edge(int(temp[0]),int(temp[1]))
        for line in edge_list_new:
            temp = line.strip().split()
            G2.add_edge(int(temp[0]), int(temp[1]))

        # The current time slice and the total number of nodes of the previous time slice are related to the two sets.
        total_nodes = set(G1.nodes()) | set(G2.nodes())

        nodes_added = set(G2.nodes()) - set(G1.nodes())
        # print ('Add the node set to：',nodes_added)
        nodes_removed = set(G1.nodes()) - set(G2.nodes())
        # print ('Delete the node set as：',nodes_removed)

        edges_added = set(G2.edges()) - set(G1.edges())
        # print ('Add the edge set to：',edges_added)
        edges_removed = set(G1.edges()) - set(G2.edges())
        # print ('Delete the side set as：',edges_removed)

        all_change_comm = set()
        # Add node processing #############################################################
        addn_ch_comm, addn_pro_edges, addn_commu = node_addition(G2, nodes_added, comm)

        edges_added = edges_added - addn_pro_edges  # Remove the processed edges.

        all_change_comm = all_change_comm | addn_ch_comm

        # Delete node processing #############################################################

        deln_ch_comm, deln_pro_edges, deln_commu = node_deletion(G1, nodes_removed, addn_commu)
        all_change_comm = all_change_comm | deln_ch_comm
        edges_removed = edges_removed - deln_pro_edges

        # Add edge processing #############################################################
        #    print('edges_added',edges_added)
        adde_ch_comm = edge_addition(edges_added, deln_commu)
        all_change_comm = all_change_comm | adde_ch_comm
        #    print('all_change_comm',all_change_comm)
        # Delete and process #############################################################
        dele_ch_comm = edge_deletion(edges_removed, deln_commu)
        all_change_comm = all_change_comm | dele_ch_comm
        #    print('all_change_comm',all_change_comm)
        unchangecomm = ()  # Unchanged community tags
        newcomm = {}  # The format is {node: community}
        newcomm = deln_commu  # Add and delete edges, only processed on existing nodes, no new nodes, delete nodes (pre-processed)
        unchangecomm = set(newcomm.values()) - all_change_comm
        unchcommunity = {key: value for key, value in newcomm.items() if
                         value in unchangecomm}  # Unchanged community: labels and nodes
        # Find the subgraph corresponding to the changing community, and then use Matthew effect dynamics to find the new community structure for the subgraph, and add the unchanged community structure to get the new community structure.
        #    print('change community:',all_change_comm)
        Gtemp = nx.Graph()
        Gtemp = getchangegraph(all_change_comm, newcomm, G2)

        unchagecom_maxlabe = 0
        if len(unchangecomm) > 0:
            unchagecom_maxlabe = max(unchangecomm)
        #    print('subG',Gtemp.edges())
        if Gtemp.number_of_edges() < 1:  # The community has not changed.
            comm = newcomm
        else:
            getnewcomm = CDME(Gtemp)

            #        print('newcomm',getnewcomm)
            # Merge the community structure, the unchanged plus the newly acquired
            mergecomm = {}  # Merge the yellow format as {node: community}
            mergecomm.update(unchcommunity)
            mergecomm.update(getnewcomm)
            # mergecomm=dict(**unchcommunity, **getnewcomm )
            comm = mergecomm  # Take the currently acquired community structure as the next community input.
            detectcom = list(conver_comm_to_lab(comm).values())
            commu_num = len(detectcom)
            tru_num = len(comm_new)
        #    print detectcom
        nodenumber = G2.number_of_nodes()
        print('T'+str(i)+' Time Film\'s Network Community Structure C'+str(i)+'*********************************************')
        # print(unchcommunity)
        # print(newcomm)
        # print(comm)
        drawcommunity(G2,comm,'./data/pic/' + pathfile + '_community_'+str(i)+'.png')
        # print ('getcommunity:',conver_comm_to_lab(comm))
        NMI, ARI, Purity, Errors = getscore(comm_new, detectcom,
                                            nodenumber)  # The first reference data is the real community, and the second is the detected community.
        #    print('community number:', len(set(comm.values())))
        #    print(comm)
        G1.clear()
        G1.add_edges_from(G2.edges())
        G2.clear()
        #    f = open(path_score,'a+')        #Write to the file
        #    f.write(path+'_'+str(i)+"\t"+str(NMI)+"\t"+str(ARI)+"\t"+str(Purity)+'\t'+str(commu_num)+'\t'+str(tru_num)+str(Errors)+"\n" )
        #    f.close()
        row = [str(i + 1), str(NMI), str(ARI), str(Purity), str(commu_num), str(tru_num), str(Errors)]
        ws.append(row)
    wb.save(path_score)
print('all done')

# edge1=set()
# edge2=set()
# edge1.add((1,2))
# edge1.add((1,3))
# edge1.add((1,4))
# edge2.add((1,2))
# edge2.add((1,3))
# edge2.add((2,3))
# print edge1
# print edge2
# print edge1-edge2
# print edge2-edge1
